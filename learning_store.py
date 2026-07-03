import hashlib
import json
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

try:
    import pymysql
except ImportError:
    pymysql = None


DATA_DIR = Path(os.environ.get("LEARNING_DATA_DIR", "data"))
EXCEL_PATH = DATA_DIR / "qa_learning_records.xlsx"
LEARNED_QA_PATH = DATA_DIR / "learned_qa.jsonl"

EXCEL_HEADERS = ["created_at", "question", "answer", "source", "status"]
LEARNING_SOURCE_MODES = {"rag_answer", "general_model_chat"}
BAD_ANSWER_MARKERS = [
    "目前售后知识库中没有相关信息",
    "暂时处理失败",
    "我不能给出具体结论",
]


def normalize_question(question):
    text = question.strip().lower()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[。！？!?,，~～.；;：:]", "", text)
    return text


def question_hash(question):
    normalized = normalize_question(question)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def ensure_excel_file():
    ensure_data_dir()
    if EXCEL_PATH.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "qa_records"
    sheet.append(EXCEL_HEADERS)
    workbook.save(EXCEL_PATH)


def append_to_excel(question, answer, source, status):
    ensure_excel_file()
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active
    sheet.append([
        datetime.now().isoformat(timespec="seconds"),
        question,
        answer,
        source,
        status,
    ])
    workbook.save(EXCEL_PATH)


def mysql_enabled():
    return os.environ.get("LEARNING_MYSQL_ENABLED", "0") == "1"


def mysql_config():
    return {
        "host": os.environ.get("MYSQL_HOST", "127.0.0.1"),
        "port": int(os.environ.get("MYSQL_PORT", "3306")),
        "user": os.environ.get("MYSQL_USER", "root"),
        "password": os.environ.get("MYSQL_PASSWORD", ""),
        "database": os.environ.get("MYSQL_DATABASE", "xiaoxiaoyi_chatbot"),
        "charset": "utf8mb4",
        "autocommit": True,
    }


def mysql_connect_without_db():
    config = mysql_config()
    database = config.pop("database")
    return pymysql.connect(**config), database


def ensure_mysql_table():
    if not mysql_enabled() or pymysql is None:
        return None

    connection = None
    try:
        connection, database = mysql_connect_without_db()
        with connection.cursor() as cursor:
            cursor.execute(
                f"CREATE DATABASE IF NOT EXISTS `{database}` "
                "DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
            )
            cursor.execute(f"USE `{database}`")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS qa_learning_records (
                    id BIGINT AUTO_INCREMENT PRIMARY KEY,
                    question_hash CHAR(64) NOT NULL,
                    question TEXT NOT NULL,
                    answer TEXT NOT NULL,
                    source VARCHAR(64) NOT NULL,
                    status VARCHAR(32) NOT NULL,
                    created_at DATETIME NOT NULL,
                    learned_at DATETIME NULL,
                    INDEX idx_question_hash (question_hash),
                    INDEX idx_status (status)
                ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
                """
            )
        return True
    except Exception:
        return False
    finally:
        if connection:
            connection.close()


def append_to_mysql(question, answer, source, status):
    if not mysql_enabled() or pymysql is None:
        return False

    connection = None
    try:
        if ensure_mysql_table() is not True:
            return False

        config = mysql_config()
        connection = pymysql.connect(**config)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO qa_learning_records
                    (question_hash, question, answer, source, status, created_at, learned_at)
                VALUES (%s, %s, %s, %s, %s, NOW(), %s)
                """,
                (
                    question_hash(question),
                    question,
                    answer,
                    source,
                    status,
                    datetime.now() if status == "learned" else None,
                ),
            )
        return True
    except Exception:
        return False
    finally:
        if connection:
            connection.close()


def delete_learned_from_mysql(question):
    if not mysql_enabled() or pymysql is None:
        return False

    connection = None
    try:
        config = mysql_config()
        connection = pymysql.connect(**config)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                DELETE FROM qa_learning_records
                WHERE question_hash = %s AND status = 'learned'
                """,
                (question_hash(question),),
            )
        return True
    except Exception:
        return False
    finally:
        if connection:
            connection.close()


def delete_learned_from_excel(question):
    if not EXCEL_PATH.exists():
        return

    workbook = load_workbook(EXCEL_PATH)
    old_sheet = workbook.active
    rows = list(old_sheet.iter_rows(values_only=True))
    if not rows:
        return

    header, data_rows = rows[0], rows[1:]
    normalized = normalize_question(question)
    kept_rows = [header]
    for row in data_rows:
        row_question = row[1] if len(row) > 1 else ""
        row_status = row[4] if len(row) > 4 else ""
        if row_status == "learned" and normalize_question(str(row_question)) == normalized:
            continue
        kept_rows.append(row)

    workbook.remove(old_sheet)
    new_sheet = workbook.create_sheet("qa_records")
    for row in kept_rows:
        new_sheet.append(row)
    workbook.save(EXCEL_PATH)


def load_learned_answers():
    if not LEARNED_QA_PATH.exists():
        return {}

    learned = {}
    with LEARNED_QA_PATH.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            learned[record.get("question_hash")] = record
    return learned


def find_learned_answer(question):
    record = load_learned_answers().get(question_hash(question))
    if not record:
        return None
    return record.get("answer")


def should_auto_learn(answer, source):
    if source not in LEARNING_SOURCE_MODES:
        return False
    if not answer or len(answer.strip()) < 8:
        return False
    return not any(marker in answer for marker in BAD_ANSWER_MARKERS)


def save_learned_answer(question, answer, source):
    ensure_data_dir()
    records = load_learned_answers()
    key = question_hash(question)
    records[key] = {
        "question_hash": key,
        "question": question,
        "answer": answer,
        "source": source,
        "learned_at": datetime.now().isoformat(timespec="seconds"),
    }

    with LEARNED_QA_PATH.open("w", encoding="utf-8") as file:
        for record in records.values():
            file.write(json.dumps(record, ensure_ascii=False) + "\n")


def record_qa(question, answer, source):
    status = "learned" if should_auto_learn(answer, source) else "recorded"
    append_to_excel(question, answer, source, status)
    append_to_mysql(question, answer, source, status)

    if status == "learned":
        save_learned_answer(question, answer, source)
        if os.environ.get("LEARNING_DELETE_AFTER_LEARNED", "0") == "1":
            delete_learned_from_excel(question)
            delete_learned_from_mysql(question)

    return status
